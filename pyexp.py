#!/usr/bin/env python
# -*- coding: utf-8 -*-

import os
import json
import openpyxl
import sys
import html

from googletrans import Translator
from json2html import *
from collections import defaultdict
from bs4 import BeautifulSoup



# translator = Translator()

def save_as_xlsx(file, name):
	wb = openpyxl.Workbook()
	ws = wb.active
	y = file
	header = []
	case2 = False;
	for n in y:
		for nn in n:
			if len(nn) == 1:
				case2 = True;
				break
			if nn in header and not case2:
				pass
			else:
				header.append(nn)
	if case2:
		for n in y:
			if n in header:
				pass
			else:
				header.append(n)
		for n in range(len(header)):
			ws.cell(row=1, column=(n + 1), value=header[n])
		for n in range(len(y)):
			ws.cell(row=2, column=(n + 1), value=str(y[header[n]]))
	else:
		for n in range(len(header)):
			ws.cell(row=1, column=(n + 1), value=header[n])
		for n in range(len(y)):
			for nn in range(len(header)):
				ws.cell(row=(n + 2), column=(nn + 1), value=str(y[n][header[nn]]))
	wb.save(filename=name + '.xlsx')

directory = 'jp_data/'

jsons = {}

itemsNames = []
achievementsNames = []
craftRequirements = []

reprNamesText = {'items':{'desc':'item_text', 
					  'item_effect_desc_text':'item_text',
					 },
			 'costume':{'desc':'costume_text'},
			 'achievements':{'desc':'achievements_text'},
			 'attack_data':{'attack_name':'master_attackbase_text'},
			 'weapons':{'desc':'weapon_text'},
			 'enemyparams':{'name_id':'enemyparam_text'}
			}

enumToDict = []

# with open(directory + 'SB_CBT_enum_dump.h', 'r') as enums:
# 	for line in enums:
# 		if 'enum class' in line:
# 			newLine = line.replace('enum class ', '').replace('{','= {')
# 			enumToDict.append(newLine)
# 		elif '=' in line:
# 			preLine = line.strip().replace(',','').split('=')
# 			newLine = '\t{0} : \'{1}\','.format(preLine[1].replace(' ', ''), preLine[0].replace(' ', ''))
# 			enumToDict.append(newLine + '\n')
# 		elif '};' in line:
# 			enumToDict.append('}\n')
# 		else:
# 			enumToDict.append('\n')

# with open(directory + 'SB_CBT_enum_dump.json', 'w') as jsonDump:
# 	jsonDump.writelines(enumToDict)

for file in os.listdir(directory):
	if file.endswith(".json"):
		with open(directory + file, 'r') as readable:
			jsons[file.split('.')[0]] = json.load(readable)


style = """<meta http-equiv="Cache-Control" content="no-cache, no-store, must-revalidate">
<meta http-equiv="Pragma" content="no-cache">
<meta http-equiv="Expires" content="0">

<style>
table.greyGridTable {
  border: 2px solid #FFFFFF;
  width: 100%;
  text-align: center;
  border-collapse: collapse;
}
table.greyGridTable td, table.greyGridTable th {
  border: 1px solid #FFFFFF;
  padding: 3px 4px;
}
table.greyGridTable tbody td {
  font-size: 15px;
}
table.greyGridTable td:nth-child(even) {
  background: #EBEBEB;
}
table.greyGridTable thead {
  background: #FFFFFF;
  border-bottom: 4px solid #333333;
}
table.greyGridTable thead th {
  font-size: 15px;
  font-weight: bold;
  color: #333333;
  text-align: center;
  border-left: 2px solid #333333;
}
table.greyGridTable thead th:first-child {
  border-left: none;
}

table.greyGridTable tfoot {
  font-size: 17px;
  font-weight: bold;
  color: #333333;
  border-top: 4px solid #333333;
}
table.greyGridTable tfoot td {
  font-size: 17px;
}

thead th {
    position: sticky;
    position: -webkit-sticky;
    top: 0;
    background: white;
    z-index: 10;
}

details summary::-webkit-details-marker {
  display: none;
}

</style>\n
"""

def getQuestStringById(long_id, id):
	for x in jsons['texts_locale_us_EN']:
		if x['name'] == questTextByType[long_id[0]]:
			for y in x['texts']:
				if id == y['id']:
					return y['text']
	return ''

def getItemStringById(id):
	for x in jsons['texts_locale_us_EN']:
		if x['name'] == 'item_text':
			for y in x['texts']:
				if id == y['id']:
					return y['text']


def getEnemyStringById(id):
	for x in jsons['texts_locale_us_EN']:
		if x['name'] == reprNamesText['enemyparams']['name_id']:
			for y in x['texts']:
				if id == y['id']:
					return y['text']

def getEnemyById(id):
	for enemy in jsons['enemyparams']:
		if enemy['enemy_id'] == id:
			return enemy

def getWeaponStringById(id):
	for x in jsons['texts_locale_us_EN']:
		if x['name'] == 'weapon_text':
			for y in x['texts']:
				if id == y['id']:
					return y['text']

def getImagineStringById(id):
	for x in jsons['texts_locale_us_EN']:
		if x['name'] == 'master_imagine_text':
			for y in x['texts']:
				if id == y['id']:
					return y['text']

def getSkillStringById(id):
	for x in jsons['texts_locale_us_EN']:
		if x['name'] == 'master_skill_data_text':
			for y in x['texts']:
				if id == y['id']:
					return y['text']

def getSkillTypeStringById(id):
	for x in jsons['texts_locale_us_EN']:
		if x['name'] == 'master_skilltype_text':
			for y in x['texts']:
				if id == y['id']:
					return y['text']

def getWeaponStringById(id):
	for x in jsons['texts_locale_us_EN']:
		if x['name'] == 'weapon_text':
			for y in x['texts']:
				if id == y['id']:
					return y['text']
def getItemById(id):
	for item in jsons['items']:
		if item['id'] == id:
			return item
	for weapon in jsons['weapons']:
		if weapon['id'] == id:
			return weapon

def getWeaponById(id):
	for weapon in jsons['weapons']:
		if weapon['id'] == id:
			return weapon

def getImagineById(id):
	for imagine in jsons['imagine']:
		if imagine['id'] == id:
			return imagine

def getTreasureById(id):
	for treasure in jsons['treasures']:
		if id == treasure['id']:
			return treasure

def getQuestByLongId(longId):
	for quest in jsons['quests']:
		if quest['long_id'] == longId:
			return quest

def getMapById(id):
	for mapItem in jsons['maps']['data']:
		if mapItem['id'] == id:
			return mapItem

def getRecipeByOutItemId(id):
	for recipe in jsons['craft']:
		if recipe['out_item_id'] == id:
			return recipe
	return "None"

def getSkillById(id):
	for skill in jsons['skill_data']:
		if skill['skill_id'] == id:
			return skill

def getLevelDataByType(type):
	return [level for level in jsons['leveldata'] if level['type'] == type]

# for item in jsons['items']:
# 	for x in jsons['texts_locale_us_EN']:
# 		if x['name'] == reprNamesText['items']['desc']:
# 			for y in x['texts']:
# 				if item['desc'] == y['id']:
# 					item['desc'] = y['text']
# 				if item['unidentified_name'] == y['id']:
# 					item['unidentified_name'] = y['text']
# 				if item['unidentified_desc'] == y['id']:
# 					item['unidentified_desc'] = y['text']
# 				if item['name'] == y['id']:
# 					item['name'] = y['text']
# 				if item['item_effect_desc_text'] == y['id']:
# 					item['item_effect_desc_text'] = y['text']
# 	itemsNames.append({
# 		'id': item['id'], 
# 		'name': item['name'], 
# 		'desc' : item['desc'], 
# 		'item_effect_desc_text' : item['item_effect_desc_text'],
# 		'unidentified_name' : item['unidentified_name'], 
# 		'unidentified_desc' : item['unidentified_desc']
# 		})


	return 'No String'

EnemiesPage = []

enemiesStruct = defaultdict(list)

# for enemyParam in jsons['enemyparams']:
# 	enemiesStruct[enemyParam['chara_bp_id']].append(enemyParam)

# for name, group in enemiesStruct.items():
# 	enemyPage = []
# 	addCommonInfo = True
# 	enemyBpID = ''
# 	for enemy in group:
# 		enemyPageItem = {}
# 		if addCommonInfo:
# 			enemyPageItem['Name'] = '<b>{0}</b>'.format(getEnemyStringById(enemy['name_id']))
# 			enemyPageItem['Race'] = enemy['race_id']
# 		enemyPageItem['Variation ID'] = enemy['enemy_id']
# 		levelParams = enemy['level_params'][0]
# 		enemyPageItem['Stats'] = 'HP: {hp_base} <i>+ {hp_factor}</i><br>Atk: {atk_base} <i>+ {atk_factor}</i><br>Def: {def_base} <i>+ {def_factor}</i><br> Sta: {sta_base}<br> EXP: {exp_base} <i>+ {exp_factor}</i><br>'.format(
# 			hp_base=levelParams['hit_point'], hp_factor=levelParams['hit_point_factor'],
# 			atk_base=levelParams['attack_power'], atk_factor=levelParams['attack_power_factor'],
# 			def_base=levelParams['defence_power'], def_factor=levelParams['defence_power_factor'],
# 			sta_base=levelParams['stamina'],
# 			exp_base=levelParams['exp'], exp_factor=levelParams['exp_factor'])
# 		enemyDrops = []
# 		for drop in enemy['drop_items']:
# 			if drop['type'] == 0:
# 				enemyDrops.append('<a href="../items/{itemID}.json">{itemName}</a> - {dropRate}% Lv{lvlMin}-{lvlMax}'.format(
# 					itemID=drop['item_index'],
# 					itemName=getItemStringById(drop['item_index']),
# 					dropRate=drop['drop_rate']/100,
# 					lvlMin=drop['level_min'],
# 					lvlMax=drop['level_max']))
# 			elif drop['type'] == 1:
# 				enemyDrops.append('L{0}-{1} - {2}%'.format(drop['drop_num_min'], drop['drop_num_max'], drop['drop_rate']/100))
# 			elif drop['type'] == 2:
# 				treasureDrop = getTreasureById(drop['item_index'])
# 				itemDrops = []
# 				rarityRateExists = False
# 				rarityRate = []
# 				if len(treasureDrop['rarity_rate']) > 1:
# 					for rate in treasureDrop['rarity_rate']:
# 						rarityRate.append(int(rate['rate'])/100)
# 						rarityRateExists = True
# 				for index, lotItem in enumerate(treasureDrop['lot_rate']):
# 					if lotItem['reward_type'] == 0:
# 						if rarityRateExists:
# 							try:
# 								rate = rarityRate[index]
# 							except:
# 								rate = float(lotItem['rate']/100)
# 						else:
# 							rate = float(lotItem['rate']/100)
# 						itemDrops.append('L{min}-{max} - {dropRate}%'.format(min=lotItem['reward_amount_min'],
# 							max=lotItem['reward_amount_max'], dropRate=str(rate)))
# 					elif lotItem['reward_type'] == 1:
# 						if rarityRateExists:
# 							try:
# 								rate = rarityRate[index]
# 							except:
# 								rate = float(lotItem['rate']/100)
# 						else:
# 							rate = float(lotItem['rate']/100)
# 						itemDrops.append('&ensp;<a href="../items/{itemID}.json">{itemName}</a> ({amountMin}-{amountMax}) R{rarity} - {dropRate}%'.format(
# 							itemID=lotItem['reward_master_id'], itemName=getItemStringById(lotItem['reward_master_id']),
# 							amountMin=lotItem['reward_amount_min'], amountMax=lotItem['reward_amount_max'],
# 							rarity=lotItem['rarity_max'], dropRate=str(rate)))
# 				enemyDrops.append('Treasure: - {dropRate}%<br>{itemDropsString}'.format(
# 					dropRate=drop['drop_rate']/100,
# 					itemDropsString='<br>'.join(itemDrops)))
# 		enemyPageItem['Drops'] = '<br>'.join(enemyDrops)
# 		enemyPageItem['Boss'] = bool(enemy['is_boss'])
# 		enemyPageItem['Skills'] = '<br>'.join([x['action_name'] for x in enemy['skill_params']])
# 		enemyPageItem['Resist Rate'] = enemy['resist_rate']
# 		enemyPageItem['Resist Dot'] = enemy['resist_dot']
# 		enemyPage.append(enemyPageItem)
# 		enemyBpID = enemy['chara_bp_id']

# 	enemyLanding = []
# 	columnsAmount = len(enemyPage)+1
# 	lablesText = [k for k, v in enemyPage[0].items()]
# 	enemyPlaceholder = defaultdict(list)
# 	for text in lablesText:
# 		for preparedEnemy in enemyPage:
# 			enemyPlaceholder[text].append(preparedEnemy[text])
# 	VariationIDToUse = enemyPlaceholder['Variation ID'].copy()
# 	lablesText.remove('Variation ID')
# 	lablesTextToUse = lablesText.copy()
# 	for text in lablesText:
# 		dataList = enemyPlaceholder[text]
# 		columnData = {}
# 		columnData['Label'] = lablesTextToUse.pop(0)
# 		for index, x in enumerate(dataList):
# 			colName = VariationIDToUse[index]
# 			columnData[colName] = dataList.pop(0)
# 		enemyLanding.append(columnData)

# 	with open('enemieshtml/' + enemyBpID + '.html', 'w', errors='ignore') as enemyPageTestFile:
# 		# enemyPageTestFile.write(json2html.convert(json=enemyLanding, table_attributes="border=\"1\" id=\"info-table\" class=\" table-condensed table table-bordered table-hover\"", escape=False))
# 		soup = BeautifulSoup(json2html.convert(json=enemyLanding, table_attributes="border=\"1\" id=\"info-table\" class=\" table-condensed table table-bordered table-hover\"", escape=False), "lxml")
# 		table = soup.find('table')
# 		table_body = table.find('tbody')
# 		for index, row in enumerate(table_body.find_all('tr')): # second row is 0
# 			if index in (0, 1):
# 				columns = row.find_all('td')
# 				for el in columns[2:]:
# 					el.decompose()
# 				columns[1]['colspan']=len(columns)
# 				columns[1]['align']='center'
# 		enemyPageTestFile.write(str(soup))
# 	with open('enemies/' + enemyBpID + '.json', 'w', errors='ignore') as enemyPageTestJson:
# 		json.dump(enemyLanding, enemyPageTestJson, indent=4)

weaponsPage = []

for weapon in jsons['weapons']:
	ESBAttribute = {
	0 : 'None',
	1 : 'Physics',
	2 : 'Earth',
	3 : 'Fire',
	4 : 'Ice',
	5 : 'Thunder',
	6 : 'Light',
	7 : 'Darkness'
	}

	ESBClassType = {
	0 : 'None',
	1 : 'ShadowAvatar',
	2 : 'SwordSlayer',
	3 : 'Artist',
	4 : 'Gunslinger',
	5 : 'Splash',
	6 : 'Blaster',
	7 : 'Berserker',
	8 : 'Engineer',
	9 : 'Striker',
	10 : 'Slayer',
	11 : 'Magician',
	12 : 'Rebellion',
	13 : 'ShortRanger',
	14 : 'LongRanger',
	15 : 'ImagineTrans',
	16 : 'Vanguard',
	17 : 'Sentinel',
	18 : 'HeavyArm',
	19 : 'Max'
	}

	if os.path.exists('E:\\BLUEPROTOCOL\\umodel\\UmodelExport\\Game\\UI\\Icon\\Class\\UI_IconClass_{:02d}'.format(weapon['equip_class']) + '.png'):
		classIcon = '<img src="images/Class/UI_IconClass_{:02d}.png" alt="None">'.format(weapon['equip_class'])
	else:
		classIcon = ESBClassType[weapon['equip_class']]

	if os.path.exists('images/Weapon/{0}.png'.format(weapon['icon_name'])):
		weaponIcon = '<img src="images/Weapon/{0}.png" alt="None">'.format(weapon['icon_name'])
	else:
		weaponIcon = ''

	recipeItem = getRecipeByOutItemId(weapon['id'])
	if recipeItem != 'None':
		obtainableString = 'Craftable from:<br>'
		for material in recipeItem['materials']:
			obtainableString += '<a href="items/{itemId}.html">{itemName}</a> <b>x{amount}</b><br>'.format(
				itemId=material['item_id'], itemName= getItemStringById(getItemById(material['item_id'])['name']),
				amount=material['need_num'])
		obtainableString += '{0} <b>L</b>'.format(recipeItem['use_money'])
	else:
		obtainableString = 'From Shop_Weapons'

	if ESBAttribute[weapon['attribute']] != 'None':
		elementalDmg = '<b>+{0} {1}</b>'.format(weapon['attribute_damage'],ESBAttribute[weapon['attribute']])
	else:
		elementalDmg = ''

	weaponLevels = []
	for idx, level in enumerate(getLevelDataByType(weapon['level_table'])):
		weaponLevels.append({
			'Lvl' : level['level'],
			'AP' : '<b>+{0}</b>'.format(weapon['enhance_status_ap']['level_' + str(idx)]),
			'EXP' : level['exp']
			})

	weaponLevels = json2html.convert(json=weaponLevels, table_attributes="border=\"1\"", escape=False)

	weaponsPage.append({
		'Class' : classIcon,
		'ID' : weapon['id'],
		'Icon' : weaponIcon,
		'Name' : '<p><b>{0}</b></p>'.format(getWeaponStringById(weapon['name'])),
		'Description' : getWeaponStringById(weapon['desc']),
		'Obtainable by' : obtainableString,
		'Elemental Additional Damage' : elementalDmg,
		'Level Table' : '<details><summary>Show</summary>{0}</details>'.format(weaponLevels),
		'Price' : 'Costs:<br><b>{0}</b><br>Sell:<br><b>{1}</b>'.format(weapon['price'], weapon['selling_price']),
		'Active' : weapon['active']
		})

with open('weaponsPage.json', 'w', encoding='utf-8') as weaponsPageJsonFile:
	json.dump(weaponsPage, weaponsPageJsonFile, indent=4)

with open('weaponsPage.html', 'w', encoding='utf-16') as weaponsPageFile:
	weaponsPageFile.write(style)
	weaponsPageFile.write(json2html.convert(json=weaponsPage, table_attributes="border=\"1\" id=\"info-table\" class=\"greyGridTable table-condensed table table-bordered table-hover\"", escape=False))

skillsPage = []

for skill in jsons['skill_data']:
	ESBClassType = {
	0 : 'None',
	1 : 'ShadowAvatar',
	2 : 'SwordSlayer',
	3 : 'Artist',
	4 : 'Gunslinger',
	5 : 'Splash',
	6 : 'Blaster',
	7 : 'Berserker',
	8 : 'Engineer',
	9 : 'Striker',
	10 : 'Slayer',
	11 : 'Magician',
	12 : 'Rebellion',
	13 : 'ShortRanger',
	14 : 'LongRanger',
	15 : 'ImagineTrans',
	16 : 'Vanguard',
	17 : 'Sentinel',
	18 : 'HeavyArm',
	19 : 'Max'
	}

	ESBSkillType = {
	-1 : 'Unknown',
	0 : 'Main Attack',
	1 : 'Sub',
	2 : 'Skill1',
	3 : 'Skill2',
	4 : 'Skill3',
	5 : 'Skill4',
	6 : 'Special',
	7 : 'Dodge',
	8 : 'Pasive1',
	9 : 'Pasive2',
	10 : 'Max'
	}

	hideSkillFromProduction = False

	if len(skill['skill_mastery_param']) > 0:
		skillLevels = []
		for level in skill['skill_mastery_param']:
			skillLevels.append('<b>{0}-</b> {1}SP - Required Lv.{2}'.format(level['level'], level['use_skill_point'], level['condition_class_level']))
		skillLevels = '<br>'.join(skillLevels)
	else:
		skillLevels = ''

	if skill['condition_skill_id'] != 0:
		unlockConditions = '<a href="#{2}"> <b>Lv.{0} {1}</b></a>'.format(skill['condition_skill_level'], getSkillStringById(getSkillById(skill['condition_skill_id'])['skill_name']), skill['condition_skill_id'])
	else:
		unlockConditions = ''
	if os.path.exists('E:\\BLUEPROTOCOL\\umodel\\UmodelExport\\Game\\UI\\Icon\\Class\\UI_IconClass_{:02d}'.format(skill['class_type']) + '.png'):
		classIcon = '<img src=images/Class/UI_IconClass_{:02d}.png alt="None">'.format(skill['class_type'])
	else:
		classIcon = ESBClassType[skill['class_type']]
		hideSkillFromProduction = True
	if 'Pasive' not in ESBSkillType[skill['skill_type']]:
		if os.path.exists('images\\Skills\\UI_PlayerSkill_{0}'.format(skill['skill_id']) + '.png'):
			skillIcon = '<img src=images/Skills/UI_PlayerSkill_{0}.png alt="None">'.format(skill['skill_id'])
		else:
			skillIcon = ''
	else:
		if os.path.exists('images\\Skills\\UI_PlayerSkillP_{0}'.format(skill['skill_id']) + '.png'):
			skillIcon = '<img src=images/Skills/UI_PlayerSkillP_{0}.png alt="None">'.format(skill['skill_id'])
		else:
			skillIcon = ''

	skillDesc = '※' + '<br><br>※'.join([ getSkillStringById(desc['desc']) for desc in skill['skill_desc_array'] ])

	if 'T skills' in skillDesc:
		hideSkillFromProduction = True

	if hideSkillFromProduction:
		continue

	skillsPage.append({
		'ID' : '<b>{0}</b>'.format(skill['skill_id']),
		'Name' : '<p id="{0}"><b>{1}</b></p>'.format(skill['skill_id'],getSkillStringById(skill['skill_name'])),
		'Icon' : skillIcon,
		'Class' : classIcon,
		'Type' : ESBSkillType[skill['skill_type']],
		'Description' : skillDesc,
		'Levels' : skillLevels,
		'Unlocks after' : unlockConditions
		})

with open('skillsPage.json', 'w', encoding='utf-8') as skillsPageJsonFile:
	json.dump(skillsPage, skillsPageJsonFile, indent=4)

with open('skillsPage.html', 'w', encoding='utf-16') as skillsPageFile:
	skillsPageFile.write(style)
	skillsPageFile.write(json2html.convert(json=skillsPage, table_attributes="border=\"1\" id=\"info-table\" class=\"greyGridTable table-condensed table table-bordered table-hover\"", escape=False))

questTextByType = {'M' : 'quest_main_text', 'S' : 'quest_sub_text', 'C' : 'quest_sub_text', 'T' : 'quest_test_text'}



questsPage = []

for quest in jsons['quests']:
	if quest['long_id'].startswith('M'):
		questType = "<img src=images/Quest/UI_IconQuest_MainPop.png>"
	elif quest['long_id'].startswith('S'):
		questType = "<img src=images/Quest/UI_IconQuest_SubPop.png>"
	elif quest['long_id'].startswith('C'):
		questType = "<img src=images/Quest/UI_IconQuest_ClassSubPop.png>"
	elif quest['long_id'].startswith('T'):
		questType = "<img src=images/Quest/UI_IconQuest_SubPop.png>"
	questSteps = ""
	for step in quest['quest_condition_steps']:
		conditions = step['condition_items'][0]
		try:
			try:
				questDesc = getQuestStringById(quest['long_id'], conditions['descriptionId'])
			except:
				questDesc = "No description"
			"""
			EQuestConditionType::Unknown = 0
			EQuestConditionType::TalkNpc = 1
			EQuestConditionType::KillEnemy = 2
			EQuestConditionType::CollectItem = 3
			EQuestConditionType::ClearID = 4
			EQuestConditionType::ArriveAt = 5
			EQuestConditionType::PayedMoney = 6
			EQuestConditionType::Emote = 7
			EQuestConditionType::WarpPortal = 8
			EQuestConditionType::Interaction = 9
			EQuestConditionType::CraftItem = 10
			EQuestConditionType::AdventurerRankUp = 11
			EQuestConditionType::AnyCondition = 12
			EQuestConditionType::CollectTriggerByEnemy = 13
			EQuestConditionType::CollectTriggerByGather = 14
			EQuestConditionType::CollectTrigger = 15
			EQuestConditionType::Num = 16
			"""
			if conditions['type'] == 1:
				questSteps += 'Step {0}: ({1}), {2}<br>'.format(step['id'][-1], "Talk to NPC", questDesc)
			elif conditions['type'] == 2:
				enemy = getEnemyById(conditions['enemyId'])
				if not enemy:
					enemy = getEnemyById(conditions['enemyId'] + '_00000')
					if not enemy:
						enemy = getEnemyById(conditions['enemyId'] + '_00_00000')
				questSteps += 'Step {stepId}: Kill {amount} <a href="enemies/{charaId}.html">{name}</a><br>'.format(stepId=conditions['id'], amount=conditions['amount'], charaId=enemy['chara_bp_id'], name=getEnemyStringById(enemy['name_id']))
			elif conditions['type'] == 3:
				if conditions['item_type'] == 1: 
					item = getItemById(conditions['item_id'])
					itemName = '<a href="items/{itemId}.html">{itemName}</a>'.format(itemId=item['id'], itemName=getItemStringById(item['name']))
				elif conditions['item_type'] == 2:
					item = getWeaponById(conditions['item_id'])
					itemName = '<a href="weapons/{itemId}.html">{itemName}</a>'.format(itemId=item['id'], itemName=getWeaponStringById(item['name']))
				elif conditions['item_type'] == 5:
					item = getImagineById(conditions['item_id'])
					itemName = '<a href="imagines/{itemId}.html">{itemName}</a>'.format(itemId=item['id'], itemName=getImagineStringById(item['imagine_name']))
				questSteps += 'Step {0}: ({1}), {2}<br>'.format(step['id'][-1], 'Get {amount} of {itemLink}'.format(amount=conditions['amount'], itemLink=itemName), questDesc.replace('{item_id}','<u>{0}</u>'.format(itemName)))
			elif conditions['type'] == 4:
				questSteps += 'Step {0}: ({1}) {2}<br>'.format(step['id'][-1], "Clear dungeon", '<img src="images/Dungeon/{iconId}.png" height="150">'.format(iconId=getMapById(conditions['dungeonId'])['menu_image']))
			elif conditions['type'] == 5:
				questSteps += 'Step {0}: ({1}), {2}<br>'.format(step['id'][-1], "Go to destination", questDesc)
			elif conditions['type'] == 6: # unused
				questSteps += 'Step {0}: ({1}), {2}<br>'.format(step['id'][-1], "Unclear", questDesc)
			elif conditions['type'] == 7:
				questSteps += 'Step {0}: ({1}), {2}<br>'.format(step['id'][-1], "Perform action", questDesc)
			elif conditions['type'] == 8:
				questSteps += 'Step {0}: ({1}), {2}<br>'.format(step['id'][-1], "Use the warp portal", questDesc)
			elif conditions['type'] == 9: # unused 
				questSteps += 'Step {0}: ({1}), {2}<br>'.format(step['id'][-1], "Interact", questDesc)
			elif conditions['type'] == 10:
				questSteps += 'Step {0}: ({1}), {2}<br>'.format(step['id'][-1], "Craft Weapon", questDesc)
			elif conditions['type'] == 11:
				questSteps += 'Step {0}: ({1}), {2}<br>'.format(step['id'][-1], "Increase adventurer rank", questDesc)
			elif conditions['type'] == 12: #unused
				questSteps += 'Step {0}: ({1}), {2}<br>'.format(step['id'][-1], "Any condition", questDesc)
			elif conditions['type'] == 13:
				enemy = getEnemyById(conditions['enemyId'])
				questSteps += 'Step {stepId}: Kill {amount} <a href="enemies/{charaId}.html">{name}</a><br>'.format(stepId=conditions['id'], amount=conditions['amount'], charaId=enemy['chara_bp_id'], name=getEnemyStringById(enemy['name_id']))
			elif conditions['type'] == 14:
				questSteps += 'Step {0}: ({1}), {2}<br>'.format(step['id'][-1], "Collect <b>{amount}</b> of <u>{itemName}</u> at {locId}".format(amount=conditions['amount'], itemName=getQuestStringById(quest['long_id'], conditions['trigger_name']), locId=conditions['locationId']), questDesc)
			elif conditions['type'] == 15:
				questSteps += 'Step {0}: ({1}), {2}<br>'.format(step['id'][-1], "Craft Weapon", questDesc)
			elif conditions['type'] == 16:
				questSteps += 'Step {0}: ({1}), {2}<br>'.format(step['id'][-1], "Craft Weapon", questDesc)
		except KeyError as e:
				questSteps += 'Step {0}: Something went wrong<br>'.format(step['id'][-1])

	questRewards = ""
	if quest['quest_rewards']:
		for reward in quest['quest_rewards']:
			if reward['type'] == 1:
				questRewards += '<b>{amount} Luno</b><br>'.format(amount=reward['amount'])
			if reward['type'] == 2:
				questRewards += '<b>{amount} Engram</b><br>'.format(amount=reward['amount'])
			if reward['type'] == 3:
				questRewards += '<b>{amount} BP</b><br>'.format(amount=reward['amount'])
			if reward['type'] == 4:
				item = getItemById(reward['item_id'])
				itemName = '<a href="items/{itemId}.html">{itemName}</a>'.format(itemId=item['id'], itemName=getItemStringById(item['name']))
				questRewards += '<b>{amount}</b> of {itemName}<br>'.format(itemName=itemName, amount=reward['amount'])
			if reward['type'] == 5:
				item = getWeaponById(reward['item_id'])
				itemName = '<a href="weapons/{itemId}.html">{itemName}</a>'.format(itemId=item['id'], itemName=getWeaponStringById(item['name']))
				questRewards += '<b>{amount}</b> of {itemName}<br>'.format(itemName=itemName, amount=reward['amount'])
			if reward['type'] == 6:
				questRewards += '<b>{amount} Costume</b><br>'.format(amount=reward['amount'])
			if reward['type'] == 7:
				item = getImagineById(reward['item_id'])
				itemName = '<a href="imagines/{itemId}.html">{itemName}</a>'.format(itemId=item['id'], itemName=getImagineStringById(item['imagine_name']))
				questRewards += '<b>{amount}</b> of {itemName}<br>'.format(itemName=itemName, amount=reward['amount'])
			if reward['type'] == 8:
				questRewards += '<b>{amount} EXP</b><br>'.format(amount=reward['amount'])
			if reward['type'] == 9:
				questRewards += '<b>{amount} Money</b><br>'.format(amount=reward['amount'])
			if reward['type'] == 10:
				questRewards += '<b>{amount} Money</b><br>'.format(amount=reward['amount'])
			if reward['type'] == 11:
				questRewards += '<b>{amount} Imagine Mount</b><br>'.format(amount=reward['amount'])
			if reward['type'] == 12:
				questRewards += '<b>{amount} Money</b><br>'.format(amount=reward['amount'])
	else:
		questRewards = "None"
	questUnlockConditions = ""
	if quest['quest_unlock_data']:
		for condition in quest['quest_unlock_data']['quest_conditions']:
			try:
				questItem = getQuestByLongId(condition['quest_id'])
				questName = getQuestStringById(questItem['long_id'], questItem['name'])
				questUnlockConditions += '<b>{questName}</b><br>'.format(questName=questName)
			except:
				questUnlockConditions += '{0}<br>'.format(condition['scenario_flag_key'])
	questsPage.append({
		'Type' : questType,
		'ID' : quest['long_id'],
		'Name' : '<b>{0}</b>'.format(getQuestStringById(quest['long_id'], quest['name'])),
		'Description' : '<br>\u25c6'.join(getQuestStringById(quest['long_id'], quest['desc']).replace('\n', '<br>').split('\u25c6')),
		#'Condition Detail' : getQuestStringById(quest['long_id'], quest['condition_detail']),
		'Steps' : questSteps,
		'Rewards' : questRewards,
		'Unlock Conditions' : questUnlockConditions
	})

with open('questsPage.json', 'w', encoding='utf-8') as questsPageJsonFile:
	json.dump(questsPage, questsPageJsonFile, indent=4)

with open('questsPage.html', 'w', encoding='utf-16') as questsPageFile:
	questsPageFile.write(style)
	questsPageFile.write(json2html.convert(json=questsPage, table_attributes="border=\"1\" id=\"info-table\" class=\"greyGridTable table-condensed table table-bordered table-hover\"", escape=False))

itemsPage = []

for item in jsons['items']:

	itemCategories = {
	0 : 'Consumable',
	1 : 'For Sale',
	2 : 'Refinement Resource',
	3 : 'Craft Resource'
	}

	boolToEng = {
	True : 'Yes',
	False : 'No'
	}

	itemCategoryString = itemCategories[item['category']]

	if os.path.exists('E:\\BLUEPROTOCOL\\umodel\\UmodelExport\\Game\\UI\\Icon\\Item\\' + item['icon_name'] + '.png'):
		icon = '<img src=images/Item/{img}.png alt="None">'.format(img=item['icon_name'])
	else:
		icon = ''

	if item['obtaining_route'] == 1:
		if len(item['obtaining_route_detail_id']) > 0:
			routeDetail = item['obtaining_route_detail_id'].split(',')
			enemyItem = getEnemyById(routeDetail[1])
			obtainableString = 'Killing <a href="enemies/{enemyId}.html">{enemyName}</a> at {locId}(Lv.{lvl})'.format(
				enemyId=enemyItem['enemy_id'], enemyName=getEnemyStringById(enemyItem['name_id']), locId=routeDetail[0],
				lvl=routeDetail[2])
		else:
			obtainableString = 'Not obtainable'
	elif item['obtaining_route'] == 2:
		if item['obtaining_route_detail_id'] != '':
			obtainableString = 'at {locId}'.format(locId=item['obtaining_route_detail_id'])
		else:
			obtainableString = ''
	elif item['obtaining_route'] == 3: 
		"""
		Material has item_type
		Recipe has use_money
		"""
		recipeItem = getRecipeByOutItemId(item['id'])
		if recipeItem != 'None':
			obtainableString = 'Craftable from:<br>'
			for material in recipeItem['materials']:
				obtainableString += '<a href="items/{itemId}.html">{itemName}</a> <b>x{amount}</b><br>'.format(
					itemId=material['item_id'], itemName= getItemStringById(getItemById(material['item_id'])['name']),
					amount=material['need_num'])
			obtainableString += '{0} <b>L</b>'.format(recipeItem['use_money'])
		else:
			obtainableString = 'Craftable, but no recipe found'
	elif item['obtaining_route'] == 4:
		obtainableString = 'Clearing dungeon <img src="images/Dungeon/{iconId}.png" height="150">'.format(iconId=getMapById(item['obtaining_route_detail_id'])['menu_image'])
	else:
		obtainableString = ''

	if bool(item['can_use']):
		if len(item['efficacies']) > 0:
			itemEfficacy = item['efficacies'][0]
			if itemEfficacy['type'] == 1:
				itemEfficacyString = 'Restores {0} HP'.format(itemEfficacy['value'])
			elif itemEfficacy['type'] == 6:
				itemEfficacyString = 'Boosts damage for {0} seconds'.format(itemEfficacy['time'])
			elif itemEfficacy['type'] == 7:
				itemEfficacyString = 'Boosts defence for {0} seconds'.format(itemEfficacy['time'])
			elif itemEfficacy['type'] == 8:
				itemEfficacyString = 'Removes fire debuff'
			elif itemEfficacy['type'] == 9:
				itemEfficacyString = 'Removes ice debuff'
			elif itemEfficacy['type'] == 10:
				itemEfficacyString = 'Removes thunder debuff'
			elif itemEfficacy['type'] == 14:
				itemEfficacyString = 'ChallengeBox'
			else:
				itemEfficacyString = 'Not implemented'
		else:
			itemEfficacyString = getItemStringById(item['item_effect_desc_text'])
	else:
		itemEfficacyString = ''
	
	if bool(item['can_use']):
		usableIcon = '<img src="images/usable.png">'
	else:
		usableIcon = '<img src="images/notusable.png">'

	if bool(item['no_sale_flag']):
		sellableIcon = '<img src="images/notsellable.png">'
	else:
		sellableIcon = '<img src="images/sellable.png">'

	itemsPage.append({'Icon' : icon,
				'ID' : '{0}'.format(item['id']),
				'Name' : '<a href="items/{0}.html"><b>{1}</b></a>'.format(item['id'], getItemStringById(item['name'])),
				'Category' : itemCategoryString, 
				'Description' :  getItemStringById(item['desc']),
				'Obtainable by' : obtainableString,
				'Item Effect' : itemEfficacyString,
				'Usable' : usableIcon,
				'Level' : item['item_level'],
				'Exp' : item['item_exp'],
				'Forbidden to sell' : sellableIcon,
				'Costs' : item['buying_price'],
				'Sold for' : item['selling_price']
				})
with open('itemsPage.html', 'w', encoding='utf-16') as itemsPageFile:
	itemsPageFile.write(style)
	itemsPageFile.write(json2html.convert(json=itemsPage, table_attributes="border=\"1\" id=\"info-table\" class=\"greyGridTable table-condensed table table-bordered table-hover\"", escape=False))

exit()

for item in jsons['items']:
	if os.path.exists('E:\\BLUEPROTOCOL\\umodel\\UmodelExport\\Game\\UI\\Icon\\Item\\' + item['icon_name'] + '.png'):
		icon = '<img src=../images/Item/{img}.png alt="None">'.format(img=item['icon_name'])
	else:
		icon = ''
	itemPage = {'Icon' : icon,
				'Name' : item['name'],
				'Description' : item['desc'],
				'Item Effect' : item['item_effect_desc_text'],
				'Usable' : boolToEng[bool(item['can_use'])],
				'Level' : item['item_level'],
				'Exp' : item['item_exp'],
				'Forbidden to sell' : boolToEng[bool(item['no_sale_flag'])],
				'Costs' : item['buying_price'],
				'Sold for' : item['selling_price'],
				'ID' : item['id']
				}
	with open('items/' + str(item['id']) + '.json', 'w') as anotherItem:
		anotherItem.seek(0)
		json.dump(itemPage, anotherItem, indent=4)
	with open('itemshtml/' + str(item['id']) + '.html', 'w', encoding='utf8') as anotherItem:
		anotherItem.write(json2html.convert(json=itemPage, table_attributes="border=\"1\" id=\"info-table\" class=\" table-condensed table table-bordered table-hover\"", escape=False))

for achievement in jsons['achievements']:
	for x in jsons['texts_locale_us_EN']:
		if x['name'] == reprNamesText['achievements']['desc']:
			for y in x['texts']:
				if achievement['desc'] == y['id']:
					achievement['desc'] = y['text']
	achievementsNames.append({
		'id' : achievement['id'],
		'desc' : achievement['desc']
		})

for weapon in jsons['weapons']:
	for x in jsons['texts_locale_us_EN']:
		if x['name'] == reprNamesText['weapons']['desc']:
			for y in x['texts']:
				if weapon['desc'] == y['id']:
					weapon['desc'] = y['text']
				if weapon['name'] == y['id']:
					weapon['name'] = y['text']


for craftItem in jsons['craft']:
	craftItem['name'] = getItemById(craftItem['out_item_id'])['name']
	for material in craftItem['materials']:
		material['name'] = getItemById(material['item_id'])['name']
	requiredItems = ','.join(['{0} x{1}'.format(material['name'], material['need_num']) for material in craftItem['materials']])
	craftRequirements.append({'id' : craftItem['id'], 'name' : craftItem['name'], 'RequiredItems': requiredItems, 'Diff' : craftItem['difficulty']})

save_as_xlsx(craftRequirements, 'craft_text')


# save_as_xlsx(achievementsNames, 'achievements_text')

# with open(directory + 'items_text.json', 'a+') as newItems:
# 	json.dump(itemsNames, newItems	, indent=4, sort_keys=True)

# save_as_xlsx(itemsNames)

# with open(directory + 'items_EN.json', 'a+') as newItems:
# 	json.dump(jsons['items'], newItems, indent=4, sort_keys=True)

